"""Nodos del pipeline de generación de reportes — FASE 7.

Produce tres salidas:

  T38  sipsa_abastecimiento_YYYYMMDD.xlsx       — 5 hojas: TD_Total, TD_Abast, TD_Destino,
                                         TD_Abast_Otros, TREF_Productos.
                                         TD_Abast incluye Proc_Part y Descr_pegar (col 9).
                                         TD_Destino incluye Ciudad_Part y Descr_pegar (col 8).
                                         Compatible con FORMATO_SIPSA_IPC.xlsm para uso manual.
  T39  Alimentos_priorizados_*.xlsx  — hoja Artículos_IPC con estructura idéntica al
                                         output del macro "PEGAR DATOS" de FORMATO_SIPSA_IPC.xlsm:
                                         col A = código IPC, cols I/K = zonas/destinos (texto),
                                         cols L-P = abastecimiento y variaciones numéricas.
  T40  historico_td_total.parquet    — acumulado mensual de TD_Total.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ─── Columnas del T38 ─────────────────────────────────────────────────────────

_COLS_TD_TOTAL = [
    "RArtículo_IPC", "Artículo_IPC",
    "AbastTotal_MesActual", "AbastTotal_MesAnterior", "AbastTotal_AnoAnterior",
    "VariacMensual", "VariacAnual",
]
# TD_Abast en T38: 9 columnas para que FORMATO_SIPSA_IPC.xlsm pueda usar VLookup en col 9
_COLS_TD_ABAST = [
    "RArtículo_IPC", "Artículo_IPC", "Departamento Proc.",
    "Sum_Ton", "Total_Artículo", "Participación",
    "Otro", "Proc_Part", "Descr_pegar",
]
# TD_Destino en T38: 8 columnas para que FORMATO_SIPSA_IPC.xlsm pueda usar VLookup en col 8
_COLS_TD_DESTINO = [
    "RArtículo_IPC", "Artículo_IPC", "Ciudad",
    "Sum_Ton", "Total_Artículo", "Participación",
    "Ciudad_Part", "Descr_pegar",
]
_COLS_TD_ABAST_OTROS = [
    "RArtículo_IPC", "Artículo_IPC", "Municipio Proc.",
    "Sum_Ton", "Total_Artículo", "Participación",
]

# ─── Índices de columna en hoja Artículos_IPC (T39) ──────────────────────────
# Estructura idéntica a FORMATO_SIPSA_IPC.xlsm post-macro "PEGAR DATOS":
#   col A (0): RArtículo_IPC (clave de búsqueda del macro VBA)
#   col B (1): Código SIPSA (vacío en el output del macro)
#   col C (2): Artículo IPC
#   col I (8): Zonas abastecedoras   ← Descr_pegar de TD_Abast
#   col K (10): Destino              ← Descr_pegar de TD_Destino
#   col L (11): Abastecimiento mes actual
#   col M (12): Abastecimiento mes anterior
#   col N (13): Abastecimiento año anterior
#   col O (14): Variación mensual (decimal, ej. -0.026)
#   col P (15): Variación anual    (decimal, ej. -0.090)
_IDX_CODIGO    = 0   # A: RArtículo_IPC
_IDX_ARTICULO  = 2   # C: Artículo IPC
_IDX_ZONAS     = 8   # I: Zonas abastecedoras
_IDX_DESTINO   = 10  # K: Destino de los alimentos
_IDX_ABAST_ACT = 11  # L: Abastecimiento mes actual
_IDX_ABAST_ANT = 12  # M: Abastecimiento mes anterior
_IDX_ABAST_A12 = 13  # N: Abastecimiento año anterior
_IDX_VARAC_M   = 14  # O: Variación mensual (numérico decimal)
_IDX_VARAC_A   = 15  # P: Variación anual   (numérico decimal)
_N_COLS        = 18

# Encabezados por defecto (estructura del FORMATO_SIPSA_IPC.xlsm → hoja Artículos_IPC)
_DEFAULT_HEADERS: list[str | None] = [
    None,
    "Código SIPSA",
    "Artículo IPC",
    "Tipo de cultivo ",
    "Tiempo de ciclo vegetativo de los cultivos",
    "Observaciones detalladas",
    "Clasificación de la observación",
    "Requerimiento hídrico por cultivo",
    "Zonas abastecedoras",
    "Porcentaje de distribución según zonas productoras en el país",
    "Destino de los alimentos (sistemas de consultas, aplicativos o similares) "
    "(Participación % por ciudad y depto)",
    "Abastecimiento total observado en el mes de actual ",
    "Abastecimiento total observado en el mes de anterior ",
    "Abastecimiento total observado en el mismo mes del año anterior  ",
    "Variacion mensual del volumen de abastecimiento",
    "Variacion anual del volumen de abastecimiento",
    "Afectaciones en la ruta",
    "Estacionalidad del consumo",
]

_MAX_ENTRADAS_TEXTO = 15

# ─── Formato visual T39 (idéntico al archivo de referencia Alimentos_priorizados) ─
_COLOR_HDR_AZUL = "FF1F497D"   # encabezados en azul oscuro (todas las cols excepto K)
_COLOR_HDR_ROJO = "FFFF0000"   # encabezado col K en rojo

_COL_WIDTHS_T39 = {
    "A": 5.7265625,  "B": 19.7265625, "C": 26.0,
    "D": 18.26953125, "E": 18.26953125, "F": 27.1796875, "G": 25.0,
    "I": 31.453125, "J": 29.26953125,
    "L": 19.26953125, "O": 18.7265625, "P": 19.54296875,
    "Q": 17.453125, "R": 18.453125,
}


def _aplicar_formato_articulos_ipc(ws, n_filas: int) -> None:
    """Aplica el formato visual de la hoja Artículos_IPC según el archivo de referencia.

    Referencia: Alimentos priorizados MAYO-26_SIPSA_20260603.xlsx
    - Fuente: Calibri 11 en toda la hoja.
    - Encabezados (fila 2): negrita, color azul (FF1F497D) excepto col K (FFFF0000).
    - Alineación centro/centro con ajuste de texto en encabezados.
    - Datos: L-N → #,##0 | O-P → 0.00% | I/K → ajuste de texto.
    - Zoom 85%, freeze en A3, alto fila 2 = 72.5 pt.
    """
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 72.5

    for col_letter, width in _COL_WIDTHS_T39.items():
        ws.column_dimensions[col_letter].width = width

    # Encabezados — fila 2 (col A tiene header None → no aplica negrita/color)
    for col_idx in range(1, _N_COLS + 1):
        cell = ws.cell(row=2, column=col_idx)
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            # Col A: header vacío; limpiar alineación que pandas aplica por defecto
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment()
        else:
            color = _COLOR_HDR_ROJO if col_letter == "K" else _COLOR_HDR_AZUL
            cell.font = Font(name="Calibri", size=11, bold=True, color=color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Datos — filas 3 en adelante
    for row_idx in range(3, n_filas + 3):
        for col_idx in range(1, _N_COLS + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_letter = get_column_letter(col_idx)
            if col_letter == "A":
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_letter == "B":
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col_letter == "C":
                cell.font = Font(name="Calibri", size=11, bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_letter in ("I", "K"):
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif col_letter in ("L", "M", "N"):
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "#,##0"
            elif col_letter in ("O", "P"):
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0.00%"
            elif col_letter == "J":
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(vertical="center")
            else:
                cell.font = Font(name="Calibri", size=11)


# ─── T38: sipsa_abastecimiento_YYYYMMDD.xlsx ─────────────────────────────────────────────

def exportar_sipsa_abastecimiento(
    td_total_variaciones: pd.DataFrame,
    td_abast_fmt: pd.DataFrame,
    td_destino_fmt: pd.DataFrame,
    td_abast_otros_fmt: pd.DataFrame,
    articulos_ipc: dict[str, Any],
    fecha_proceso: str,
    ruta_reporting: str,
) -> pd.DataFrame:
    """Genera sipsa_abastecimiento_YYYYMMDD.xlsx con 5 hojas, compatible con FORMATO_SIPSA_IPC.xlsm.

    TD_Abast incluye las columnas ``Proc_Part`` (texto por fila) y ``Descr_pegar``
    (texto completo del artículo en col 9), que el macro VBA "PEGAR DATOS" lee para
    rellenar la columna Zonas abastecedoras de la hoja Artículos_IPC.

    TD_Destino incluye ``Ciudad_Part`` y ``Descr_pegar`` (col 8) con el mismo propósito
    para la columna Destino de los alimentos.

    Args:
        td_total_variaciones: TD_Total con variaciones (F5).
        td_abast_fmt: TD_Abast formateada y ordenada (F6).
        td_destino_fmt: TD_Destino ordenada (F6).
        td_abast_otros_fmt: TD_Abast_Otros formateada y ordenada (F6).
        articulos_ipc: Diccionario ``{"codigos": ..., "variedades": ...}`` del mes.
        fecha_proceso: Cadena YYYYMMDD que forma parte del nombre del archivo.
        ruta_reporting: Directorio de destino (relativo al proyecto).

    Returns:
        DataFrame de una fila con metadatos del archivo exportado.
    """
    Path(ruta_reporting).mkdir(parents=True, exist_ok=True)
    filepath = Path(ruta_reporting) / f"SIPSA_ABASTECIMIENTO_{fecha_proceso}.xlsx"

    td_abast_xlsm   = _agregar_cols_abast(td_abast_fmt, td_abast_otros_fmt)
    td_destino_xlsm = _agregar_cols_destino(td_destino_fmt)
    tref            = _construir_tref_productos(articulos_ipc)

    hojas = {
        "TD_Total":       td_total_variaciones[_COLS_TD_TOTAL],
        "TD_Abast":       td_abast_xlsm[_COLS_TD_ABAST],
        "TD_Destino":     td_destino_xlsm[_COLS_TD_DESTINO],
        "TD_Abast_Otros": td_abast_otros_fmt[_COLS_TD_ABAST_OTROS],
    }

    with pd.ExcelWriter(str(filepath), engine="openpyxl") as writer:
        for nombre_hoja, df_hoja in hojas.items():
            df_hoja.to_excel(writer, sheet_name=nombre_hoja, index=False)
        tref.to_excel(writer, sheet_name="TREF_Productos", index=False)

    filas_total = sum(len(df) for df in hojas.values())
    log.info(
        "exportar_sipsa_abastecimiento OK | archivo=%s | hojas=%d | filas_totales=%d | tref_articulos=%d",
        filepath.name, len(hojas) + 1, filas_total, len(tref),
    )
    return pd.DataFrame([{
        "archivo": str(filepath),
        "hojas": len(hojas) + 1,
        "filas_td_total":       len(hojas["TD_Total"]),
        "filas_td_abast":       len(hojas["TD_Abast"]),
        "filas_td_destino":     len(hojas["TD_Destino"]),
        "filas_td_abast_otros": len(hojas["TD_Abast_Otros"]),
        "filas_tref":           len(tref),
    }])


def _agregar_cols_abast(
    td_abast: pd.DataFrame,
    td_otros: pd.DataFrame,
) -> pd.DataFrame:
    """Agrega Otro, Proc_Part y Descr_pegar a TD_Abast para compatibilidad con XLSM.

    - ``Proc_Part`` (col 8): texto formateado de la fila individual.
    - ``Descr_pegar`` (col 9): texto completo multilinea del artículo, igual en
      todas las filas del mismo artículo. El macro VBA usa VLookup(código, A:I, 9)
      para rellenar la columna Zonas abastecedoras de Artículos_IPC.
    """
    df = td_abast.copy()
    df["Otro"] = None

    paises_por_art = _construir_lookup_paises(td_otros)

    def _parte_abast(depto: str, p: float, cod: Any) -> str:
        if depto == "N.A.":
            return f"OTRO   {_fmt_pct(p)} "
        return f"{depto}  {_fmt_pct(p)} "

    df["Proc_Part"] = df.apply(
        lambda r: _parte_abast(r["Departamento Proc."], r["Participación"], r["RArtículo_IPC"]),
        axis=1,
    )

    # Descr_pegar: texto completo por artículo (máx _MAX_ENTRADAS_TEXTO)
    descr_por_art: dict[Any, str] = {}
    for cod, grp in df.groupby("RArtículo_IPC", sort=True):
        partes = [
            _parte_abast(r["Departamento Proc."], r["Participación"], cod)
            for _, r in grp.head(_MAX_ENTRADAS_TEXTO).iterrows()
        ]
        descr_por_art[cod] = "\n".join(partes)

    df["Descr_pegar"] = df["RArtículo_IPC"].map(descr_por_art)
    return df


def _agregar_cols_destino(td_destino: pd.DataFrame) -> pd.DataFrame:
    """Agrega Ciudad_Part y Descr_pegar a TD_Destino para compatibilidad con XLSM.

    - ``Ciudad_Part`` (col 7): texto formateado de la fila individual.
    - ``Descr_pegar`` (col 8): texto completo multilinea del artículo. El macro VBA
      usa VLookup(código, A:H, 8) para rellenar la columna Destino de Artículos_IPC.
    """
    df = td_destino.copy()

    df["Ciudad_Part"] = df.apply(
        lambda r: f"{r['Ciudad']}  {_fmt_pct(r['Participación'])}",
        axis=1,
    )

    descr_por_art: dict[Any, str] = {}
    for cod, grp in df.groupby("RArtículo_IPC", sort=True):
        partes = [
            f"{r['Ciudad']}  {_fmt_pct(r['Participación'])}"
            for _, r in grp.head(_MAX_ENTRADAS_TEXTO).iterrows()
        ]
        descr_por_art[cod] = "\n".join(partes)

    df["Descr_pegar"] = df["RArtículo_IPC"].map(descr_por_art)
    return df


def _construir_tref_productos(articulos_ipc: dict[str, Any]) -> pd.DataFrame:
    """Construye TREF_Productos: lista de artículos IPC con sus códigos del mes."""
    codigos = articulos_ipc.get("codigos", {})
    filas = [
        {"Código SIPSA": int(cod), "Artículo IPC": str(art)}
        for art, cod in sorted(codigos.items(), key=lambda x: int(x[1]))
    ]
    return pd.DataFrame(filas, columns=["Código SIPSA", "Artículo IPC"])


# ─── T39: Alimentos_priorizados_*.xlsx ────────────────────────────────────────

def exportar_alimentos_priorizados(
    td_total_variaciones: pd.DataFrame,
    td_abast_fmt: pd.DataFrame,
    td_destino_fmt: pd.DataFrame,
    td_abast_otros_fmt: pd.DataFrame,
    mes_actual_nombre: str,
    anio_actual: int,
    fecha_proceso: str,
    ruta_reporting: str,
    archivo_entrada: str = "",
) -> pd.DataFrame:
    """Genera Alimentos_priorizados_MES-AA_SIPSA_YYYYMMDD.xlsx.

    Replica exactamente el output del macro "PEGAR DATOS" de FORMATO_SIPSA_IPC.xlsm:

    - Carga la plantilla Artículos_IPC del archivo de entrada (si existe).
    - Para cada artículo en la plantilla, busca sus datos en las tablas TD y los
      escribe en las columnas correspondientes.
    - Col A  (0): código IPC (RArtículo_IPC).
    - Col I  (8): Zonas abastecedoras (texto multilinea con departamentos y %).
    - Col K (10): Destino de los alimentos (texto multilinea con ciudades y %).
    - Cols L-N (11-13): Abastecimiento actual, anterior y del año anterior.
    - Cols O-P (14-15): Variación mensual y anual como decimal (ej. -0,026).
    - Fila 1 vacía, fila 2 encabezados, datos desde fila 3 (igual que el XLSM).

    Args:
        td_total_variaciones: TD_Total con variaciones (F5).
        td_abast_fmt: TD_Abast formateada (F6).
        td_destino_fmt: TD_Destino formateada (F6).
        td_abast_otros_fmt: TD_Abast_Otros formateada (F6).
        mes_actual_nombre: Nombre español del mes actual (ej: ``"Enero"``).
        anio_actual: Año del período actual.
        fecha_proceso: Cadena YYYYMMDD para el nombre del archivo.
        ruta_reporting: Directorio de destino.
        archivo_entrada: Ruta al archivo de entrada con la plantilla Artículos_IPC.
            Si está vacío o el archivo no existe, se usan encabezados por defecto.

    Returns:
        DataFrame de una fila con metadatos del archivo exportado.
    """
    Path(ruta_reporting).mkdir(parents=True, exist_ok=True)
    anio_corto = str(anio_actual)[-2:]
    mes_corto  = mes_actual_nombre[:3].lower()
    nombre     = f"Alimentos_priorizados_{mes_corto}{anio_corto}_SIPSA_{fecha_proceso}.xlsx"
    filepath   = Path(ruta_reporting) / nombre

    headers, template_rows = _cargar_template_articulos(archivo_entrada)
    resumen_lookup = _construir_resumen_lookup(
        td_total_variaciones, td_abast_fmt, td_destino_fmt, td_abast_otros_fmt
    )
    output_rows = _rellenar_template(template_rows, resumen_lookup, headers)
    df_output   = pd.DataFrame(output_rows, columns=headers)

    with pd.ExcelWriter(str(filepath), engine="openpyxl") as writer:
        df_output.to_excel(
            writer, sheet_name="Artículos_IPC", index=False, startrow=1,
        )
        _aplicar_formato_articulos_ipc(writer.sheets["Artículos_IPC"], len(df_output))

    n_articulos = len(df_output)
    log.info(
        "exportar_alimentos_priorizados OK | archivo=%s | articulos=%d",
        nombre, n_articulos,
    )
    return pd.DataFrame([{"archivo": str(filepath), "articulos": n_articulos}])


def _cargar_template_articulos(
    archivo_entrada: str,
) -> tuple[list[str | None], list[list[Any]]]:
    """Carga la hoja Artículos_IPC del archivo de entrada como plantilla.

    Returns:
        headers: lista de 18 encabezados (fila 1 de la hoja, 0-indexed).
        rows: lista de listas, una por artículo, con los 18 valores de cada fila.
    """
    if not archivo_entrada:
        return _plantilla_vacia()

    ruta = Path(archivo_entrada)
    if not ruta.is_absolute():
        ruta = Path.cwd() / archivo_entrada
    if not ruta.exists():
        log.warning(
            "_cargar_template_articulos | archivo no encontrado: '%s'; usando plantilla vacía.",
            archivo_entrada,
        )
        return _plantilla_vacia()

    raw = pd.read_excel(str(ruta), sheet_name=None, header=None)
    hoja = next(
        (s for s in raw if s.lower().startswith("art") and "ipc" in s.lower()),
        None,
    )
    if hoja is None:
        log.warning(
            "_cargar_template_articulos | hoja Artículos_IPC no encontrada en '%s'; "
            "usando plantilla vacía.",
            ruta.name,
        )
        return _plantilla_vacia()

    df = raw[hoja]
    while df.shape[1] < _N_COLS:
        df[df.shape[1]] = None

    headers: list[str | None] = df.iloc[1, :_N_COLS].tolist()
    rows: list[list[Any]] = []
    for _, fila in df.iloc[2:].iterrows():
        valores = fila.iloc[:_N_COLS].tolist()
        nombre  = valores[_IDX_ARTICULO]
        if nombre is None or (isinstance(nombre, float) and pd.isna(nombre)):
            continue
        rows.append(valores)

    log.info(
        "_cargar_template_articulos | hoja='%s' | articulos=%d",
        hoja, len(rows),
    )
    return headers, rows


def _plantilla_vacia() -> tuple[list[str | None], list[list[Any]]]:
    """Retorna encabezados por defecto y filas vacías (sin plantilla disponible)."""
    return list(_DEFAULT_HEADERS), []


def _construir_resumen_lookup(
    td_total: pd.DataFrame,
    td_abast: pd.DataFrame,
    td_destino: pd.DataFrame,
    td_otros: pd.DataFrame,
) -> dict[str, dict[str, Any]]:
    """Construye diccionario artículo_casefold → datos dinámicos del mes.

    Las variaciones se almacenan como fracciones decimales numéricas (ej. -0,026),
    igual que el output del macro "PEGAR DATOS" en las celdas de Artículos_IPC.
    """
    base = td_total[_COLS_TD_TOTAL + ["VariacMensual_num", "VariacAnual_num"]].copy()
    zonas_df    = _texto_zonas(td_abast, td_otros)
    destinos_df = _texto_destinos(td_destino)

    merged = base.merge(zonas_df,    on="RArtículo_IPC", how="left")
    merged = merged.merge(destinos_df, on="RArtículo_IPC", how="left")

    lookup: dict[str, dict[str, Any]] = {}
    for _, row in merged.iterrows():
        key = str(row["Artículo_IPC"]).strip().casefold()
        lookup[key] = {
            "nombre":       str(row["Artículo_IPC"]).strip(),
            "codigo":       int(row["RArtículo_IPC"]),
            "zonas_texto":  row.get("Zonas abastecedoras", ""),
            "destino_texto":row.get("Destino", ""),
            "abast_act":    row["AbastTotal_MesActual"],
            "abast_ant":    row["AbastTotal_MesAnterior"],
            "abast_a12":    row["AbastTotal_AnoAnterior"],
            "varac_m":      row["VariacMensual_num"] / 100,
            "varac_a":      row["VariacAnual_num"] / 100,
        }
    return lookup


def _rellenar_template(
    template_rows: list[list[Any]],
    lookup: dict[str, dict[str, Any]],
    headers: list[str | None],
) -> list[list[Any]]:
    """Rellena las columnas dinámicas en cada fila de la plantilla.

    Replica la lógica VBA de FORMATO_SIPSA_IPC.xlsm:
      - Lee el código del artículo en col A (posición 0).
      - Llena cols I, K, L, M, N, O, P con datos de las tablas TD.
    Artículos del lookup no presentes en la plantilla se agregan al final.
    """
    output: list[list[Any]] = []
    articulos_en_template: set[str] = set()

    for fila in template_rows:
        nombre = fila[_IDX_ARTICULO]
        key    = str(nombre).strip().casefold() if nombre else ""
        articulos_en_template.add(key)

        datos     = lookup.get(key, {})
        nueva_fila = list(fila)

        if datos:
            nueva_fila[_IDX_CODIGO]    = datos["codigo"]
            nueva_fila[_IDX_ZONAS]     = datos["zonas_texto"]
            nueva_fila[_IDX_DESTINO]   = datos["destino_texto"]
            nueva_fila[_IDX_ABAST_ACT] = datos["abast_act"]
            nueva_fila[_IDX_ABAST_ANT] = datos["abast_ant"]
            nueva_fila[_IDX_ABAST_A12] = datos["abast_a12"]
            nueva_fila[_IDX_VARAC_M]   = datos["varac_m"]
            nueva_fila[_IDX_VARAC_A]   = datos["varac_a"]
        else:
            log.warning(
                "_rellenar_template | artículo '%s' en plantilla sin datos calculados.",
                nombre,
            )

        output.append(nueva_fila)

    for key, datos in sorted(lookup.items(), key=lambda x: x[1]["codigo"]):
        if key in articulos_en_template:
            continue
        nueva_fila = [None] * _N_COLS
        nueva_fila[_IDX_CODIGO]    = datos["codigo"]
        nueva_fila[_IDX_ARTICULO]  = datos["nombre"]
        nueva_fila[_IDX_ZONAS]     = datos["zonas_texto"]
        nueva_fila[_IDX_DESTINO]   = datos["destino_texto"]
        nueva_fila[_IDX_ABAST_ACT] = datos["abast_act"]
        nueva_fila[_IDX_ABAST_ANT] = datos["abast_ant"]
        nueva_fila[_IDX_ABAST_A12] = datos["abast_a12"]
        nueva_fila[_IDX_VARAC_M]   = datos["varac_m"]
        nueva_fila[_IDX_VARAC_A]   = datos["varac_a"]
        log.warning(
            "_rellenar_template | artículo código=%d no está en la plantilla; "
            "agregado al final.",
            datos["codigo"],
        )
        output.append(nueva_fila)

    return output


def _texto_zonas(td_abast: pd.DataFrame, td_otros: pd.DataFrame) -> pd.DataFrame:
    """Genera texto multilinea de zonas abastecedoras por artículo (col I de Artículos_IPC)."""
    paises_por_art = _construir_lookup_paises(td_otros)
    filas = []
    for cod, grp in td_abast.groupby("RArtículo_IPC", sort=True):
        partes: list[str] = []
        for _, fila in grp.head(_MAX_ENTRADAS_TEXTO).iterrows():
            depto = fila["Departamento Proc."]
            p     = fila["Participación"]
            if depto == "N.A.":
                parte = f"OTRO   {_fmt_pct(p)}"
            else:
                parte = f"{depto}  {_fmt_pct(p)}"
            partes.append(parte)
        filas.append({"RArtículo_IPC": cod, "Zonas abastecedoras": "\n".join(partes)})
    return pd.DataFrame(filas)


def _texto_destinos(td_destino: pd.DataFrame) -> pd.DataFrame:
    """Genera texto multilinea de destinos de consumo por artículo (col K de Artículos_IPC)."""
    filas = []
    for cod, grp in td_destino.groupby("RArtículo_IPC", sort=True):
        partes = [
            f"{r['Ciudad']}  {_fmt_pct(r['Participación'])}"
            for _, r in grp.head(_MAX_ENTRADAS_TEXTO).iterrows()
        ]
        filas.append({"RArtículo_IPC": cod, "Destino": "\n".join(partes)})
    return pd.DataFrame(filas)


def _construir_lookup_paises(td_otros: pd.DataFrame) -> dict[int, list[str]]:
    """Dict artículo → lista de países de importación ordenados por Participación."""
    if td_otros.empty:
        return {}
    lookup: dict[int, list[str]] = {}
    for cod, grp in td_otros.groupby("RArtículo_IPC", sort=True):
        paises = (
            grp.sort_values("Participación", ascending=False)["Municipio Proc."].tolist()
        )
        lookup[int(cod)] = paises
    return lookup


def _fmt_pct(valor: float) -> str:
    """Formatea un porcentaje con coma decimal y 2 decimales. Ej: ``24,07%``."""
    return f"{valor:.2f}".replace(".", ",") + "%"


# ─── T40: Histórico mensual en Parquet ────────────────────────────────────────

def guardar_historico(
    td_total_variaciones: pd.DataFrame,
    mes_actual_nombre: str,
    anio_actual: int,
) -> pd.DataFrame:
    """Agrega el mes actual al acumulado histórico mensual de TD_Total."""
    df = td_total_variaciones[_COLS_TD_TOTAL + ["VariacMensual_num", "VariacAnual_num"]].copy()
    df["mes"]  = mes_actual_nombre
    df["anio"] = anio_actual

    log.info(
        "guardar_historico OK | mes=%s %d | articulos=%d",
        mes_actual_nombre, anio_actual, len(df),
    )
    return df

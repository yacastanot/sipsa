"""Compara el formato de T39 generado vs el archivo de referencia."""
import openpyxl
from openpyxl.utils import get_column_letter

REF = (
    r"C:\Users\Jeferson\OneDrive - Cloud Integration Hub\Documentos"
    r"\DANE Automatización\SIPSA IPC Actualizado\02Salida"
    r"\Alimentos priorizados MAYO-26_SIPSA_20260603.xlsx"
)
GEN = r"data\08_reporting\Alimentos_priorizados_may26_SIPSA_20260611.xlsx"

def rgb(c):
    if c is None: return "None"
    if c.type == "rgb": return c.rgb
    return f"({c.type}:{c.theme}/{c.tint})"

def resumen_celda(cell):
    f = cell.font
    a = cell.alignment
    return {
        "val":    repr(cell.value)[:40] if cell.value is not None else None,
        "bold":   f.bold,
        "size":   f.size,
        "font":   f.name,
        "color":  rgb(f.color),
        "h":      a.horizontal,
        "v":      a.vertical,
        "wrap":   a.wrap_text,
        "nfmt":   cell.number_format,
    }

def comparar(ref_ws, gen_ws, fila, col_idx):
    r = resumen_celda(ref_ws.cell(row=fila, column=col_idx))
    g = resumen_celda(gen_ws.cell(row=fila, column=col_idx))
    diffs = {k: (r[k], g[k]) for k in r if r[k] != g[k] and k != "val"}
    return diffs

wb_ref = openpyxl.load_workbook(REF)
wb_gen = openpyxl.load_workbook(GEN)
ws_ref = wb_ref.active
ws_gen = wb_gen.active

print(f"REF: {ws_ref.max_row} filas x {ws_ref.max_column} cols | zoom={ws_ref.sheet_view.zoomScale}")
print(f"GEN: {ws_gen.max_row} filas x {ws_gen.max_column} cols | zoom={ws_gen.sheet_view.zoomScale}")
print(f"REF freeze: {ws_ref.freeze_panes} | GEN freeze: {ws_gen.freeze_panes}")
print()

print("=== FILA 2 (encabezados) ===")
ok = True
for col_idx in range(1, 19):
    col_letter = get_column_letter(col_idx)
    diffs = comparar(ws_ref, ws_gen, 2, col_idx)
    if diffs:
        print(f"  Col {col_letter}: DIFFS = {diffs}")
        ok = False
if ok:
    print("  Todos los encabezados coinciden [OK]")

print()
print("=== FILA 3 (primera fila de datos) ===")
ok = True
for col_idx in range(1, 19):
    col_letter = get_column_letter(col_idx)
    diffs = comparar(ws_ref, ws_gen, 3, col_idx)
    ignore_val = True  # valores distintos es esperado
    if diffs:
        print(f"  Col {col_letter}: DIFFS = {diffs}")
        ok = False
if ok:
    print("  Formato de la primera fila de datos coincide [OK]")

print()
print("=== ANCHOS DE COLUMNA ===")
ref_w = {k: v.width for k, v in ws_ref.column_dimensions.items()}
gen_w = {k: v.width for k, v in ws_gen.column_dimensions.items()}
all_cols = sorted(set(ref_w) | set(gen_w))
for col_letter in all_cols:
    rv = ref_w.get(col_letter, "default")
    gv = gen_w.get(col_letter, "default")
    status = "[OK]" if rv == gv else "[DIFF]"
    print(f"  Col {col_letter}: ref={rv}  gen={gv}  {status}")

print()
print("=== ALTO FILA 2 ===")
r2h_ref = ws_ref.row_dimensions[2].height if 2 in ws_ref.row_dimensions else "default"
r2h_gen = ws_gen.row_dimensions[2].height if 2 in ws_gen.row_dimensions else "default"
print(f"  ref={r2h_ref}  gen={r2h_gen}  {'[OK]' if r2h_ref == r2h_gen else '[DIFF]'}")

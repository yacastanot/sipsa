"""Inspecciona col A (código) y col C (artículo) en el GEN y el REF."""
import openpyxl

REF = (
    r"C:\Users\Jeferson\OneDrive - Cloud Integration Hub\Documentos"
    r"\DANE Automatización\SIPSA IPC Actualizado\02Salida"
    r"\Alimentos priorizados MAYO-26_SIPSA_20260603.xlsx"
)
GEN = r"data\08_reporting\Alimentos_priorizados_may26_SIPSA_20260611.xlsx"
INP = r"data\01_raw\Alimentos priorizados may2026_SIPSA.xlsx"

print("=== GEN (archivo generado) ===")
wb = openpyxl.load_workbook(GEN, data_only=True)
ws = wb.active
print(f"{'Fila':<5} {'ColA (código)':<20} {'ColC (artículo)'}")
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    cod = row[0].value
    art = row[2].value
    if cod is not None or art is not None:
        print(f"  {row[0].row:<5} {repr(cod):<20} {repr(art)[:60]}")

print()
print("=== REF (referencia) ===")
wb_ref = openpyxl.load_workbook(REF, data_only=True)
ws_ref = wb_ref.active
print(f"{'Fila':<5} {'ColA (código)':<20} {'ColC (artículo)'}")
for row in ws_ref.iter_rows(min_row=3, max_row=min(ws_ref.max_row, 35)):
    cod = row[0].value
    art = row[2].value
    if cod is not None or art is not None:
        print(f"  {row[0].row:<5} {repr(cod):<20} {repr(art)[:60]}")

print()
print("=== INPUT template (Artículos_IPC sheet) ===")
try:
    import pandas as pd
    df = pd.read_excel(INP, sheet_name="Artículos_IPC", header=None)
    print(f"  Dims: {df.shape}")
    for i, fila in df.iloc[2:].iterrows():
        a = fila.iloc[0]  # col A
        c = fila.iloc[2]  # col C
        if pd.notna(c):
            print(f"  row={i+1:<5} colA={repr(a):<20} colC={repr(str(c))[:60]}")
except Exception as e:
    print(f"  Error: {e}")

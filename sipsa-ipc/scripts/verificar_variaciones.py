"""Verifica que las variaciones en T39 coincidan con el archivo de referencia."""
import openpyxl

REF = (
    r"C:\Users\Jeferson\OneDrive - Cloud Integration Hub\Documentos"
    r"\DANE Automatización\SIPSA IPC Actualizado\02Salida"
    r"\Alimentos priorizados MAYO-26_SIPSA_20260603.xlsx"
)
GEN = r"data\08_reporting\Alimentos_priorizados_may26_SIPSA_20260611.xlsx"

wb_ref = openpyxl.load_workbook(REF, data_only=True)
wb_gen = openpyxl.load_workbook(GEN, data_only=True)
ws_ref = wb_ref.active
ws_gen = wb_gen.active

# Construir índice código → fila en cada archivo
def indice_por_codigo(ws):
    idx = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        cod = row[0].value  # col A = código IPC
        if isinstance(cod, (int, float)) and cod:
            idx[int(cod)] = row
    return idx

ref_idx = indice_por_codigo(ws_ref)
gen_idx = indice_por_codigo(ws_gen)

codigos_comunes = sorted(set(ref_idx) & set(gen_idx))
print(f"Códigos en REF: {len(ref_idx)} | en GEN: {len(gen_idx)} | comunes: {len(codigos_comunes)}")
print()
print(f"{'Cod':<6} {'REF O':>18} {'GEN O':>22} {'REF P':>18} {'GEN P':>22} {'OK'}")
print("-" * 100)

all_ok = True
for cod in codigos_comunes:
    r_ref = ref_idx[cod]
    r_gen = gen_idx[cod]
    o_ref = r_ref[14].value   # col O (0-based 14)
    o_gen = r_gen[14].value
    p_ref = r_ref[15].value   # col P (0-based 15)
    p_gen = r_gen[15].value

    if o_ref is None or o_gen is None or p_ref is None or p_gen is None:
        status = "NULL"
        all_ok = False
    elif abs(o_ref - o_gen) < 1e-8 and abs(p_ref - p_gen) < 1e-8:
        status = "[OK]"
    else:
        status = f"[DIFF] O_err={abs(o_ref-o_gen):.2e} P_err={abs(p_ref-p_gen):.2e}"
        all_ok = False

    print(f"{cod:<6} {o_ref!r:>18} {o_gen!r:>22} {p_ref!r:>18} {p_gen!r:>22}  {status}")

print()
print("Resultado:", "TODOS OK" if all_ok else "HAY DIFERENCIAS")

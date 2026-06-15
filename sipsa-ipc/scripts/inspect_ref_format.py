"""Inspecciona el formato del archivo de referencia Alimentos_priorizados."""
import openpyxl
from openpyxl.utils import get_column_letter

ref = (
    r"C:\Users\Jeferson\OneDrive - Cloud Integration Hub\Documentos"
    r"\DANE Automatización\SIPSA IPC Actualizado\02Salida"
    r"\Alimentos priorizados MAYO-26_SIPSA_20260603.xlsx"
)

wb = openpyxl.load_workbook(ref)
ws = wb.active
print("Hoja activa:", ws.title)
print("Hojas:", wb.sheetnames)
print("Dims:", ws.dimensions)
print("Filas:", ws.max_row, "| Cols:", ws.max_column)
print()
print("Orientacion:", ws.page_setup.orientation)
print("Zoom:", ws.sheet_view.zoomScale)
print("Freeze panes:", ws.freeze_panes)
print()

def rgb(color):
    if color is None:
        return "None"
    if color.type == "rgb":
        return color.rgb
    return repr(color)

# Fila 1 (fila en blanco?)
print("=== FILA 1 ===")
for cell in ws[1]:
    if cell.value is not None:
        print(f"  Col {cell.column_letter}: val={repr(cell.value)[:60]}")

print()
print("=== FILA 2 (encabezados) ===")
for cell in ws[2]:
    if cell.value is not None:
        f = cell.font
        fi = cell.fill
        al = cell.alignment
        bd = cell.border
        print(f"  Col {cell.column_letter}: val={repr(cell.value)[:50]}")
        print(f"    font: name={f.name}, size={f.size}, bold={f.bold}, italic={f.italic}, color={rgb(f.color)}")
        print(f"    fill_type={fi.fill_type}, fg={rgb(fi.fgColor)}, bg={rgb(fi.bgColor)}")
        print(f"    align: h={al.horizontal}, v={al.vertical}, wrap={al.wrap_text}")
        print(f"    number_format={cell.number_format}")
print()

print("=== FILA 3 (primera fila de datos) ===")
for cell in ws[3]:
    f = cell.font
    fi = cell.fill
    al = cell.alignment
    print(f"  Col {cell.column_letter}: val={repr(cell.value)[:50]}")
    print(f"    font: name={f.name}, size={f.size}, bold={f.bold}, color={rgb(f.color)}")
    print(f"    fill_type={fi.fill_type}, fg={rgb(fi.fgColor)}")
    print(f"    align: h={al.horizontal}, v={al.vertical}, wrap={al.wrap_text}")
    print(f"    number_format={cell.number_format}")

print()
print("=== ANCHOS DE COLUMNA ===")
for col_letter, col_dim in sorted(ws.column_dimensions.items()):
    print(f"  Col {col_letter}: width={col_dim.width}, hidden={col_dim.hidden}")

print()
print("=== ALTOS DE FILA (primeras 5) ===")
for row_idx in range(1, 6):
    rd = ws.row_dimensions.get(row_idx)
    print(f"  Fila {row_idx}: height={rd.height if rd else 'default'}")

print()
print("=== FILA 4 (segunda fila de datos) ===")
for cell in ws[4]:
    f = cell.font
    fi = cell.fill
    print(f"  Col {cell.column_letter}: val={repr(cell.value)[:50]}")
    print(f"    font: name={f.name}, size={f.size}, bold={f.bold}")
    print(f"    fill_type={fi.fill_type}, fg={rgb(fi.fgColor)}")
    print(f"    number_format={cell.number_format}")

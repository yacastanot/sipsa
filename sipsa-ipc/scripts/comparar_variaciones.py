"""Compara columnas O y P (variaciones) entre T39 generado y archivo de referencia."""
import openpyxl

REF = (
    r"C:\Users\Jeferson\OneDrive - Cloud Integration Hub\Documentos"
    r"\DANE Automatización\SIPSA IPC Actualizado\02Salida"
    r"\Alimentos priorizados MAYO-26_SIPSA_20260603.xlsx"
)
GEN = r"data\08_reporting\Alimentos_priorizados_may26_SIPSA_20260611.xlsx"

wb_ref = openpyxl.load_workbook(REF, data_only=True)
wb_gen = openpyxl.load_workbook(GEN, data_only=True)
ws_ref = wb_ref.active
ws_gen = wb_gen.active

print(f"{'Fila':<5} {'Col':<4} {'REF valor':<30} {'REF nfmt':<15} {'GEN valor':<30} {'GEN nfmt':<15}")
print("-" * 105)

# Filas 3..32 (datos), cols O=15 y P=16
for row_idx in range(3, min(ws_ref.max_row, ws_gen.max_row) + 1):
    # Artículo desde col C (índice 3)
    art_ref = ws_ref.cell(row=row_idx, column=3).value
    art_gen = ws_gen.cell(row=row_idx, column=3).value

    for col_idx, col_name in [(15, "O"), (16, "P")]:
        c_ref = ws_ref.cell(row=row_idx, column=col_idx)
        c_gen = ws_gen.cell(row=row_idx, column=col_idx)

        v_ref = c_ref.value
        v_gen = c_gen.value
        nfmt_ref = c_ref.number_format
        nfmt_gen = c_gen.number_format

        match = "==" if v_ref == v_gen else "!="
        print(
            f"{row_idx:<5} {col_name:<4} "
            f"{repr(v_ref):<30} {nfmt_ref:<15} "
            f"{repr(v_gen):<30} {nfmt_gen:<15}  {match}"
        )

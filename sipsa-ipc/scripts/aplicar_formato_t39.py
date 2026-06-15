"""Aplica formato visual al archivo Alimentos_priorizados (T39) existente.

Uso:
    python scripts/aplicar_formato_t39.py [ruta_archivo.xlsx]

Si no se pasa argumento, busca el archivo más reciente en data/08_reporting/
"""
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

_N_COLS = 18
_COLOR_HDR_AZUL = "FF1F497D"
_COLOR_HDR_ROJO = "FFFF0000"

_COL_WIDTHS = {
    "A": 5.7265625,  "B": 19.7265625, "C": 26.0,
    "D": 18.26953125, "E": 18.26953125, "F": 27.1796875, "G": 25.0,
    "I": 31.453125, "J": 29.26953125,
    "L": 19.26953125, "O": 18.7265625, "P": 19.54296875,
    "Q": 17.453125, "R": 18.453125,
}


def aplicar_formato(filepath: Path) -> None:
    print(f"Abriendo: {filepath}")
    wb = load_workbook(str(filepath))

    hoja = next((s for s in wb.sheetnames if "art" in s.lower() and "ipc" in s.lower()), None)
    if hoja is None:
        print("ERROR: hoja Artículos_IPC no encontrada.")
        return
    ws = wb[hoja]

    n_filas = ws.max_row - 2   # descontar fila 1 (blank) y fila 2 (headers)
    print(f"  Hoja: {hoja!r} | Filas de datos: {n_filas}")

    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 72.5

    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Encabezados (fila 2) — col A tiene header None: fuente normal, limpiar alineación
    for col_idx in range(1, _N_COLS + 1):
        cell = ws.cell(row=2, column=col_idx)
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment()
        else:
            color = _COLOR_HDR_ROJO if col_letter == "K" else _COLOR_HDR_AZUL
            cell.font = Font(name="Calibri", size=11, bold=True, color=color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Datos (filas 3 en adelante)
    for row_idx in range(3, n_filas + 3):
        for col_idx in range(1, _N_COLS + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_letter = get_column_letter(col_idx)
            if col_letter == "A":
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_letter == "B":
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col_letter == "C":
                cell.font = Font(name="Calibri", size=11, bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_letter in ("I", "K"):
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif col_letter in ("L", "M", "N"):
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "#,##0"
            elif col_letter in ("O", "P"):
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0.00%"
            elif col_letter == "J":
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(vertical="center")
            else:
                cell.font = Font(name="Calibri", size=11)

    wb.save(str(filepath))
    print(f"  Formato aplicado y guardado: {filepath.name}")


def main() -> None:
    if len(sys.argv) > 1:
        ruta = Path(sys.argv[1])
    else:
        carpeta = Path("data/08_reporting")
        candidatos = sorted(carpeta.glob("Alimentos_priorizados_*.xlsx"), key=lambda p: p.stat().st_mtime)
        if not candidatos:
            print("No se encontró ningún archivo Alimentos_priorizados_*.xlsx en data/08_reporting/")
            sys.exit(1)
        ruta = candidatos[-1]

    if not ruta.exists():
        print(f"Archivo no encontrado: {ruta}")
        sys.exit(1)

    aplicar_formato(ruta)


if __name__ == "__main__":
    main()

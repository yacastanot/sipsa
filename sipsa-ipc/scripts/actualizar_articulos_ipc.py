"""
Paso 0 - Preparacion mensual SIPSA IPC.

Valida el archivo de entrada y actualiza parameters_articulos_ipc.yml antes
de ejecutar kedro run.  Debe correrse una vez al inicio de cada mes.

Uso:
    python scripts/actualizar_articulos_ipc.py

Secciones:
  [1/4] Periodos en BASE SIPSA_A     - t, t-1, t-12 presentes
  [2/4] Valores faltantes Depto/Mpio - NA/vacio/CNA -> OTRO (pipeline lo corrige)
  [3/4] Consistencia Articulos_IPC   - lista igual en ambas hojas
  [4/4] parameters_articulos_ipc.yml - articulos y variedades actualizados
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from ruamel.yaml import YAML

PROJECT_ROOT = Path(__file__).resolve().parent.parent
PARAMS_FILE   = PROJECT_ROOT / "conf/base/parameters.yml"
ARTICULOS_FILE = PROJECT_ROOT / "conf/base/parameters_articulos_ipc.yml"

HOJA_BASE      = "BASE SIPSA_A"
HOJA_ALIMENTOS = "Alimentos IPC Vs SIPSA_A"
HOJA_FORMATO   = "Articulos_IPC"

MESES_ES = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12,
}

_ADVERTENCIAS: list[str] = []
_ERRORES:      list[str] = []


def _ok(msg: str)   -> None: print(f"  [OK]   {msg}")
def _warn(msg: str) -> None:
    print(f"  [WARN] {msg}")
    _ADVERTENCIAS.append(msg)
def _err(msg: str)  -> None:
    print(f"  [ERR]  {msg}")
    _ERRORES.append(msg)


# ─── YAML ────────────────────────────────────────────────────────────────────

def _leer_yaml(ruta: Path) -> dict:
    yaml = YAML()
    with open(ruta, encoding="utf-8") as f:
        return yaml.load(f)


# ─── Archivo de entrada ──────────────────────────────────────────────────────

def _resolver_ruta(archivo: str) -> Path:
    ruta = PROJECT_ROOT / archivo
    if ruta.exists():
        return ruta
    candidatos = sorted((PROJECT_ROOT / "data/01_raw").glob("Alimentos priorizados*.xlsx"))
    if candidatos:
        _warn(f"'{Path(archivo).name}' no existe; usando '{candidatos[0].name}'.")
        return candidatos[0]
    _err(f"No se encontro el archivo de entrada: {archivo}")
    sys.exit(1)


def _verificar_hojas(ruta: Path) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(ruta, read_only=True)
    hojas = set(wb.sheetnames)
    wb.close()
    for hoja in (HOJA_BASE, HOJA_ALIMENTOS):
        if hoja in hojas:
            _ok(f"Hoja '{hoja}' encontrada.")
        else:
            _err(f"Hoja '{hoja}' NO encontrada en el archivo.")
    # Formato: nombre puede variar ligeramente (acento)
    formato_ok = any(h.lower().startswith("art") and "ipc" in h.lower() for h in hojas)
    if formato_ok:
        _ok(f"Hoja de formato 'Articulos_IPC' encontrada.")
    else:
        _warn("Hoja de formato 'Articulos_IPC' no detectada (nombre inesperado).")


# ─── [1/4] PERIODOS ──────────────────────────────────────────────────────────

def _validar_periodos(ruta: Path, params: dict) -> None:
    print("\n[1/4] PERIODOS EN BASE SIPSA_A")

    mes_actual   = params["mes_actual_nombre"].strip().lower()
    mes_anterior = params["mes_anterior_nombre"].strip().lower()
    anio_actual  = int(params["anio_actual"])
    anio_ant     = int(params["anio_anterior"])

    num_actual  = MESES_ES.get(mes_actual)
    num_anterior = MESES_ES.get(mes_anterior)

    if not num_actual or not num_anterior:
        _err(f"Nombre de mes no reconocido: '{mes_actual}' / '{mes_anterior}'")
        return

    # t-12: mismo mes del año anterior
    periodo_t   = (anio_actual, num_actual)
    periodo_tm1 = (anio_actual if num_anterior > num_actual else anio_actual - 1 if num_actual == 1 else anio_actual,
                   num_anterior)
    # Para t-1 cuando mes_actual = enero, el mes anterior es diciembre del año anterior
    if num_actual == 1:
        periodo_tm1 = (anio_actual - 1, 12)

    periodo_t12 = (anio_ant, num_actual)

    try:
        df = pd.read_excel(ruta, sheet_name=HOJA_BASE, usecols=["FechaEncuesta"],
                           parse_dates=["FechaEncuesta"])
    except Exception as e:
        _err(f"No se pudo leer BASE SIPSA_A: {e}")
        return

    periodos_data = set(
        df["FechaEncuesta"].dropna().dt.to_period("M").unique()
    )

    def _check(anio: int, mes: int, etiqueta: str) -> None:
        periodo = pd.Period(f"{anio}-{mes:02d}", freq="M")
        n = int((df["FechaEncuesta"].dt.year.eq(anio) &
                 df["FechaEncuesta"].dt.month.eq(mes)).sum())
        if periodo in periodos_data:
            _ok(f"{etiqueta} ({periodo}): {n:,} filas")
        else:
            _err(f"{etiqueta} ({periodo}): NO encontrado en FechaEncuesta")

    _check(*periodo_t,   f"t    - {params['mes_actual_nombre']} {anio_actual}")
    _check(*periodo_tm1, f"t-1  - {params['mes_anterior_nombre']} {periodo_tm1[0]}")
    _check(*periodo_t12, f"t-12 - {params['mes_actual_nombre']} {anio_ant}")


# ─── [2/4] DEPARTAMENTO / MUNICIPIO ─────────────────────────────────────────

_VARIANTES = {
    "Departamento Proc.": ["Departamento Proc.", "Departamento"],
    "Municipio Proc.":    ["Municipio Proc.", "Municipio de Colombia / País Proc."],
}
_VALORES_VACIOS = {"", "N/A", "NA", "CNA", "nan", "NAN"}


def _validar_depto_mpio(ruta: Path) -> None:
    print("\n[2/4] DEPARTAMENTO Y MUNICIPIO - VALORES FALTANTES")

    try:
        df = pd.read_excel(ruta, sheet_name=HOJA_BASE, dtype=str)
    except Exception as e:
        _err(f"No se pudo leer BASE SIPSA_A: {e}")
        return

    for nombre_estandar, variantes in _VARIANTES.items():
        col_real = next((v for v in variantes if v in df.columns), None)
        if col_real is None:
            _err(f"No se encontro columna '{nombre_estandar}' ni sus variantes.")
            continue

        if col_real != nombre_estandar:
            _warn(f"'{col_real}' se renombrara a '{nombre_estandar}' en el pipeline.")

        mask = df[col_real].isna() | df[col_real].str.strip().isin(_VALORES_VACIOS)
        n = int(mask.sum())
        if n == 0:
            _ok(f"'{nombre_estandar}': sin valores faltantes.")
        else:
            _warn(f"'{nombre_estandar}': {n:,} valores NA/vacio/CNA -> el pipeline los reemplazara con 'OTRO'.")


# Valores que corresponden a encabezados de columna en la hoja de priorizados,
# no a nombres de artículos o variedades reales (fila 1: col4="IPC", col5="SIPSA").
_CABECERAS_HOJA = {"IPC", "SIPSA"}


# ─── [3/4] CONSISTENCIA ARTICULOS_IPC ────────────────────────────────────────

def _validar_articulos_formato(ruta: Path) -> None:
    print("\n[3/4] CONSISTENCIA Articulos_IPC vs Alimentos IPC Vs SIPSA_A")

    # Leer lista de la hoja de formato
    import openpyxl
    wb = openpyxl.load_workbook(ruta, read_only=True)
    hoja_fmt = next(
        (h for h in wb.sheetnames if h.lower().startswith("art") and "ipc" in h.lower()),
        None,
    )
    if hoja_fmt is None:
        _warn("Hoja 'Articulos_IPC' no encontrada; omitiendo verificacion.")
        wb.close()
        return

    ws = wb[hoja_fmt]
    articulos_formato = set()
    for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
        val = row[2] if len(row) > 2 else None
        if not val or not str(val).strip():
            continue
        texto = str(val).strip().upper()
        # Omitir celdas que son encabezados de columna
        if "ARTÍCULO" in texto or "ARTICULO" in texto or texto == "ARTÍCULO IPC":
            continue
        articulos_formato.add(texto)
    wb.close()

    # Leer lista de la hoja de priorizados (excluir encabezados de columna)
    df_prio = pd.read_excel(ruta, sheet_name=HOJA_ALIMENTOS, header=None)
    articulos_prio = set(
        v for v in (
            df_prio.iloc[2:, 2]
            .dropna()
            .str.strip()
            .str.upper()
            .tolist()
        )
        if v not in _CABECERAS_HOJA
    )

    solo_formato = articulos_formato - articulos_prio
    solo_prio    = articulos_prio - articulos_formato

    _ok(f"Articulos_IPC: {len(articulos_formato)} | Alimentos IPC Vs SIPSA_A: {len(articulos_prio)}")

    if not solo_formato and not solo_prio:
        _ok("Listas identicas.")
    if solo_formato:
        _warn(f"En Articulos_IPC pero NO en priorizados: {', '.join(sorted(solo_formato))}")
    if solo_prio:
        _warn(f"En priorizados pero NO en Articulos_IPC: {', '.join(sorted(solo_prio))}")


# ─── [4/4] ARTICULOS IPC YML ─────────────────────────────────────────────────

def _leer_hoja_priorizados(ruta: Path) -> tuple[list[str], list[tuple[str, str]]]:
    df = pd.read_excel(ruta, sheet_name=HOJA_ALIMENTOS, header=None)
    inicio = 2

    articulos_mes = [
        v for v in (
            df.iloc[inicio:, 2].dropna().str.strip().str.upper().tolist()
        )
        if v not in _CABECERAS_HOJA
    ]

    mapping_df = df.iloc[inicio:, [4, 5]].copy()
    mapping_df.columns = ["ipc", "sipsa"]
    mapping_df = mapping_df.dropna(subset=["sipsa"])
    mapping_df["ipc"] = mapping_df["ipc"].ffill()
    mapping_df = mapping_df.dropna(subset=["ipc"])

    mapping_mes = [
        (str(r["sipsa"]).strip(), str(r["ipc"]).strip().upper())
        for _, r in mapping_df.iterrows()
        if str(r["sipsa"]).strip()
        and str(r["ipc"]).strip()
        and str(r["sipsa"]).strip().upper() not in _CABECERAS_HOJA
        and str(r["ipc"]).strip().upper() not in _CABECERAS_HOJA
    ]
    return articulos_mes, mapping_mes


def _calcular_diffs(articulos_mes, mapping_mes, codigos, variedades):
    codigos_upper  = {k.strip().upper() for k in codigos}
    variedades_cf  = {k.casefold() for k in variedades}
    set_mes = set(articulos_mes)
    nuevos     = sorted(set_mes - codigos_upper)
    eliminados = sorted(codigos_upper - set_mes)
    variedades_nuevas = [
        (sipsa, ipc) for sipsa, ipc in mapping_mes
        if sipsa.casefold() not in variedades_cf
    ]
    return nuevos, eliminados, variedades_nuevas


def _asignar_codigos(nuevos, eliminados, codigos):
    codigos_upper = {k.strip().upper(): v for k, v in codigos.items()}
    libres = sorted(codigos_upper[a] for a in eliminados if a in codigos_upper)
    max_cod = max(codigos_upper.values()) if codigos_upper else 1000
    cola = list(libres)
    asignaciones: dict[str, int] = {}
    for art in nuevos:
        asignaciones[art] = cola.pop(0) if cola else (max_cod := max_cod + 1)
    eliminados_upper = {e.upper() for e in eliminados}
    nuevo_codigos = {k: v for k, v in codigos.items()
                     if k.strip().upper() not in eliminados_upper}
    for art, cod in sorted(asignaciones.items(), key=lambda x: x[1]):
        nuevo_codigos[art] = cod
    return nuevo_codigos, asignaciones


def _guardar_articulos(nuevo_codigos: dict, nuevo_variedades: dict) -> None:
    yaml = YAML()
    yaml.preserve_quotes = True
    yaml.width = 120
    with open(ARTICULOS_FILE, encoding="utf-8") as f:
        doc = yaml.load(f)
    doc["articulos_ipc"]["codigos"]   = nuevo_codigos
    doc["articulos_ipc"]["variedades"] = nuevo_variedades
    with open(ARTICULOS_FILE, "w", encoding="utf-8") as f:
        yaml.dump(doc, f)


def _actualizar_yml(ruta: Path) -> None:
    print("\n[4/4] ACTUALIZACION parameters_articulos_ipc.yml")

    datos     = _leer_yaml(ARTICULOS_FILE)
    codigos   = dict(datos["articulos_ipc"]["codigos"])
    variedades = dict(datos["articulos_ipc"]["variedades"])

    articulos_mes, mapping_mes = _leer_hoja_priorizados(ruta)
    nuevos, eliminados, variedades_nuevas = _calcular_diffs(
        articulos_mes, mapping_mes, codigos, variedades
    )

    if not nuevos and not eliminados and not variedades_nuevas:
        _ok("Sin cambios necesarios; parameters_articulos_ipc.yml esta actualizado.")
        return

    if eliminados:
        _warn(f"Articulos que salen ({len(eliminados)}): {', '.join(eliminados)}")
    if nuevos:
        _warn(f"Articulos nuevos ({len(nuevos)}): {', '.join(nuevos)}")
    if variedades_nuevas:
        for sipsa, ipc in variedades_nuevas:
            _warn(f"Nueva variedad: '{sipsa}' -> {ipc}")

    nuevo_codigos, asignaciones = _asignar_codigos(nuevos, eliminados, codigos)
    for art, cod in sorted(asignaciones.items(), key=lambda x: x[1]):
        _ok(f"Codigo asignado: {art}: {cod}")

    nuevo_variedades = dict(variedades)
    for sipsa, ipc in variedades_nuevas:
        nuevo_variedades[sipsa] = ipc

    _guardar_articulos(nuevo_codigos, nuevo_variedades)
    _ok("parameters_articulos_ipc.yml actualizado.")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main() -> None:
    params  = _leer_yaml(PARAMS_FILE)
    archivo = params["archivo_entrada"]

    print("=" * 60)
    print("  PASO 0 - PREPARACION MENSUAL SIPSA IPC")
    print("=" * 60)
    print(f"\nArchivo : {archivo}")
    print(f"Periodo : {params['mes_actual_nombre']} {params['anio_actual']}")

    ruta = _resolver_ruta(archivo)
    print(f"\n[ARCHIVO]")
    _verificar_hojas(ruta)

    _validar_periodos(ruta, params)
    _validar_depto_mpio(ruta)
    _validar_articulos_formato(ruta)
    _actualizar_yml(ruta)

    print("\n" + "=" * 60)
    if _ERRORES:
        print(f"  RESULTADO: {len(_ERRORES)} ERROR(ES) - Corregir antes de kedro run")
        for e in _ERRORES:
            print(f"    - {e}")
        sys.exit(1)
    elif _ADVERTENCIAS:
        print(f"  RESULTADO: {len(_ADVERTENCIAS)} ADVERTENCIA(S) - Revisar si es necesario")
        print("  Puede ejecutar: kedro run")
    else:
        print("  RESULTADO: TODO OK - Puede ejecutar: kedro run")
    print("=" * 60)


if __name__ == "__main__":
    main()

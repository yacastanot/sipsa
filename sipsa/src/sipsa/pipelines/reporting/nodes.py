"""Nodos de reporte: genera el Excel formateado del boletín SIPSA.

Patrón Strategy: cada tipo de fila (_tipo) tiene su propia estrategia de renderizado.
Para añadir un nuevo tipo de fila, basta con crear una subclase de FilaStrategy
y registrarla en _ESTRATEGIAS — sin tocar generar_excel.
"""
import logging
import os
from abc import ABC, abstractmethod
from typing import Any, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ─── Estilos compartidos ──────────────────────────────────────────────────────
_FONT_NAME = "Futura Std Book"
_SZ        = 9

_FONT_BOLD   = Font(name=_FONT_NAME, size=_SZ, bold=True)
_FONT_NORMAL = Font(name=_FONT_NAME, size=_SZ, bold=False)

_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

_NUM_FMT    = "#,##0"
_COL_WIDTHS = {1: 38, 2: 12.86, 3: 13.43, 4: 11.86, 5: 9.71}


def _texto(row: pd.Series) -> str:
    val = row["Productos y mercados"]
    return str(val) if pd.notna(val) else ""


# ─── Patrón Strategy ──────────────────────────────────────────────────────────

class FilaStrategy(ABC):
    """Contrato para renderizar un tipo de fila del boletín en la hoja de cálculo."""

    @abstractmethod
    def renderizar(self, ws: Worksheet, fila: int, row: pd.Series) -> None:
        """Escribe y estiliza la fila `fila` usando los datos de `row`."""


class CuadroStrategy(FilaStrategy):
    """Título de cuadro: negrita sin relleno ni combinación de celdas."""

    def renderizar(self, ws: Worksheet, fila: int, row: pd.Series) -> None:
        celda           = ws.cell(row=fila, column=1, value=_texto(row))
        celda.font      = _FONT_BOLD
        celda.alignment = _ALIGN_LEFT


class ColumnasStrategy(FilaStrategy):
    """Cabecera de columnas: negrita, sin relleno."""

    _ENCABEZADOS = [
        "Productos y mercados",
        "Precio mínimo",
        "Precio máximo",
        "Precio medio",
        "Tendencia",
    ]

    def renderizar(self, ws: Worksheet, fila: int, row: pd.Series) -> None:
        for col_idx, texto in enumerate(self._ENCABEZADOS, start=1):
            celda           = ws.cell(row=fila, column=col_idx, value=texto)
            celda.font      = _FONT_BOLD
            celda.alignment = _ALIGN_LEFT if col_idx == 1 else _ALIGN_CENTER


class ProductoStrategy(FilaStrategy):
    """Nombre del producto: negrita sin relleno."""

    def renderizar(self, ws: Worksheet, fila: int, row: pd.Series) -> None:
        celda           = ws.cell(row=fila, column=1, value=_texto(row))
        celda.font      = _FONT_BOLD
        celda.alignment = _ALIGN_LEFT


class DatoStrategy(FilaStrategy):
    """Fila de precio por mercado: precios con formato de miles, centrados."""

    def renderizar(self, ws: Worksheet, fila: int, row: pd.Series) -> None:
        celda_text           = ws.cell(row=fila, column=1, value=_texto(row))
        celda_text.font      = _FONT_NORMAL
        celda_text.alignment = _ALIGN_LEFT

        for col_idx, valor in enumerate(
            [row["Precio mínimo"], row["Precio máximo"], row["Precio medio"]], start=2
        ):
            celda = ws.cell(row=fila, column=col_idx)
            if pd.notna(valor):
                try:
                    celda.value         = int(round(float(valor)))
                    celda.number_format = _NUM_FMT
                except (TypeError, ValueError):
                    celda.value = valor
            celda.font      = _FONT_NORMAL
            celda.alignment = _ALIGN_CENTER

        tend       = row["Tendencia"]
        celda_tend = ws.cell(
            row=fila, column=5,
            value=tend if pd.notna(tend) and str(tend) != "nan" else "",
        )
        celda_tend.font      = _FONT_NORMAL
        celda_tend.alignment = _ALIGN_CENTER


class SeparadorStrategy(FilaStrategy):
    """Fila en blanco de 6 pt entre productos."""

    def renderizar(self, ws: Worksheet, fila: int, row: pd.Series) -> None:
        ws.row_dimensions[fila].height = 6


class DefaultStrategy(FilaStrategy):
    """Fallback: escribe el texto sin estilo adicional."""

    def renderizar(self, ws: Worksheet, fila: int, row: pd.Series) -> None:
        ws.cell(row=fila, column=1, value=_texto(row))


# Registro de estrategias: _tipo → instancia (singletons, sin estado mutable)
_ESTRATEGIAS: Dict[str, FilaStrategy] = {
    "cuadro":    CuadroStrategy(),
    "columnas":  ColumnasStrategy(),
    "producto":  ProductoStrategy(),
    "dato":      DatoStrategy(),
    "separador": SeparadorStrategy(),
}
_ESTRATEGIA_DEFAULT = DefaultStrategy()


# ─── Nodo Kedro ───────────────────────────────────────────────────────────────

def generar_excel(
    boletin_filas: pd.DataFrame,
    fecha: str,
    ruta_salida: str,
) -> Dict[str, Any]:
    """Escribe el boletín formateado en un archivo Excel.

    Args:
        boletin_filas: DataFrame con filas del boletín y columna _tipo.
        fecha: Fecha del boletín (ej. "27FEB2026"), usada en el nombre del archivo.
        ruta_salida: Directorio donde se guarda el Excel de salida.

    Returns:
        Dict con la ruta del archivo generado y el número de filas escritas.
    """
    os.makedirs(ruta_salida, exist_ok=True)
    nombre_archivo = f"Boletin_{fecha}_python.xlsx"
    ruta_archivo   = os.path.join(ruta_salida, nombre_archivo)

    wb       = Workbook()
    ws       = wb.active
    ws.title = "Print 1 - Conjunto de datos"

    for col_idx, ancho in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    fila_excel = 1
    for _, row in boletin_filas.iterrows():
        estrategia = _ESTRATEGIAS.get(row["_tipo"], _ESTRATEGIA_DEFAULT)
        estrategia.renderizar(ws, fila_excel, row)
        fila_excel += 1

    wb.save(ruta_archivo)
    logger.info("Excel guardado en: %s (%d filas escritas)", ruta_archivo, fila_excel - 1)
    return {"ruta_archivo": ruta_archivo, "filas_escritas": fila_excel - 1}
